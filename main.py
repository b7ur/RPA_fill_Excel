#encoding=utf-8
import tkinter as tk
from tkinter import filedialog, messagebox
import shutil
from tkinter import ttk
from openpyxl import Workbook
# from paddleocr import PaddleOCR, draw_ocr  # 导入PaddleOCR和绘图函数
from openpyxl import load_workbook
import subprocess
# import openpyxl
import os
import win32com.client

def convert_and_read_xls():
    # 弹出对话框选择 xls 文件
    data=[]
    xls_file_path = filedialog.askopenfilename(
        title='选择一个 .xls 文件',
        filetypes=[('Excel files', '*.xls')]
    )

    if not xls_file_path:
        return  # 如果没有选择文件，就不做任何操作

    # 定义 xls 转换为 xlsx 的文件路径
    xlsx_file_path = os.path.splitext(xls_file_path)[0] + '.xlsx'

    # 使用 win32com 打开 xls 文件并另存为 xlsx
    excel = win32com.client.gencache.EnsureDispatch("Excel.Application")
    wb = excel.Workbooks.Open(os.path.abspath(xls_file_path))
    wb.SaveAs(os.path.abspath(xlsx_file_path), FileFormat=51)  # 51 对应 xlsx 格式
    wb.Close()
    excel.Quit()

    # 使用 openpyxl 读取 xlsx 文件
    wb = load_workbook(xlsx_file_path)
    ws = wb.active

    # 将 xlsx 文件中的数据追加到 data 数组
    #     with open('input.txt', 'w', encoding='utf-8') as file:
    # for row in ws.iter_rows(values_only=True):
    
    #         for cell_value in row:
    #         # 将 cell_value 写入文件，并在每个 cell_value 后面添加换行符
    #             file.write(str(cell_value) + '\n')
    #             data.append(cell_value)
        
    for row in ws.iter_rows(values_only=True):
        for cell_value in row:
            data.append(cell_value)

    with open('input.txt', 'w', encoding='utf-8') as file:
        for i in range(1,len(data)):
        # 将 cell_value 写入文件，并在每个 cell_value 后面添加换行符
            file.write(str(i) + "\t"+str(data[i]) + '\n')
    #             data.append(cell_value)
    print("输出input.txt成功")         

    return data


def conv_dict():
    input_dict = {}
    data_dict = {}

    with open('input.txt', 'r', encoding='utf-8') as file_input:
        for line_number, line in enumerate(file_input, start=1):  # 假设第一行是标题，从第二行开始读取
            parts = line.strip().split()
            if len(parts) >= 2:  # 确保行中有足够的数据
                input_dict[line_number] = parts[1]  # 假设第二列的数据是我们想要的

    # 读取data.txt文件并替换第二列
    with open('data.txt', 'r', encoding='utf-8') as file_data, \
        open('data_replaced.txt', 'w', encoding='utf-8') as file_output:  # 将替换结果写入新文件
        for line in file_data:
            parts = line.strip().split(';')
            if len(parts) == 2:
                col_identifier, num_str = parts
                try:
                    # 尝试从input_dict中获取对应的值
                    replacement = input_dict[int(num_str)]
                    # 写入新的data_dict字典
                    data_dict[col_identifier] = replacement
                    # 写入替换后的行到新文件
                    file_output.write(f"{col_identifier};{replacement}\n")
                except KeyError:
                    # 如果索引超出范围，则弹出消息框提示
                    tk.Tk().withdraw()  # 不显示主窗口
                    messagebox.showerror("索引超出范围", "data.txt中的某个索引在input.txt中不存在。")
                    break
                except ValueError:
                    # 如果转换为整数失败，则弹出消息框提示
                    tk.Tk().withdraw()  # 不显示主窗口
                    messagebox.showerror("值错误", "data.txt中的数字格式不正确。")
                    break
            else:
                # 如果data.txt文件的行格式不正确，则写入原始行到新文件
                file_output.write(line)
    return data_dict


def extr():
    data=convert_and_read_xls()
    
    data_dict=conv_dict()
    

    wb_new = load_workbook("template.xlsx")
    ws_new = wb_new.active
    
    for key,val in data_dict.items():
        print(key,val)
        ws_new[key]=val
        wb_new.save("output.xlsx")
        wb_new.close()
    tk.messagebox.showinfo("完成", "新文件output.xlsx")

    return data_dict
        

root = tk.Tk()
root.title("读取并转换 Excel 文件")

# 创建一个按钮，点击时会触发 convert_and_read_xls 函数
read_button = tk.Button(root, text="读取xls文件", command=extr)
read_button.pack()

# 设置窗口大小
root.geometry('300x150')

# 启动 Tkinter 事件循环
root.mainloop()